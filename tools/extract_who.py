"""Extract the -2 SD, median and +2 SD curves from official WHO XLSX tables."""

from pathlib import Path
import csv
from urllib.request import urlretrieve
import openpyxl


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / ".tmp-who"
OUTPUT = ROOT / "app" / "src" / "main" / "assets" / "who"

TABLES = {
    "lhfa-girls.xlsx": (
        "height_girl.csv",
        "https://www.who.int/tools/child-growth-standards/standards/length-height-for-age",
        "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators/length-height-for-age/expandable-tables/lhfa-girls-zscore-expanded-tables.xlsx?sfvrsn=27f1e2cb_10",
    ),
    "lhfa-boys.xlsx": (
        "height_boy.csv",
        "https://www.who.int/tools/child-growth-standards/standards/length-height-for-age",
        "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators/length-height-for-age/expandable-tables/lhfa-boys-zscore-expanded-tables.xlsx?sfvrsn=7b4a3428_12",
    ),
    "wfa-girls.xlsx": (
        "weight_girl.csv",
        "https://www.who.int/tools/child-growth-standards/standards/weight-for-age",
        "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators/weight-for-age/expanded-tables/wfa-girls-zscore-expanded-tables.xlsx?sfvrsn=f01bc813_10",
    ),
    "wfa-boys.xlsx": (
        "weight_boy.csv",
        "https://www.who.int/tools/child-growth-standards/standards/weight-for-age",
        "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators/weight-for-age/expanded-tables/wfa-boys-zscore-expanded-tables.xlsx?sfvrsn=65cce121_10",
    ),
}


def extract(source: Path, target: Path, source_url: str) -> None:
    sheet = openpyxl.load_workbook(source, read_only=True, data_only=True).active
    headers = {cell.value: index for index, cell in enumerate(next(sheet.iter_rows()), start=1)}
    required = ["Day", "SD2neg", "SD0", "SD2"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"{source.name}: missing columns {missing}")

    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        day = values[headers["Day"] - 1]
        if day is None:
            continue
        rows.append(
            (
                int(day),
                float(values[headers["SD2neg"] - 1]),
                float(values[headers["SD0"] - 1]),
                float(values[headers["SD2"] - 1]),
            )
        )

    if rows[0][0] != 0 or rows[-1][0] < 1825:
        raise ValueError(f"{source.name}: unexpected day range {rows[0][0]}..{rows[-1][0]}")
    if any(current[0] != previous[0] + 1 for previous, current in zip(rows, rows[1:])):
        raise ValueError(f"{source.name}: non-contiguous day range")

    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", newline="", encoding="utf-8") as handle:
        handle.write(f"# WHO Child Growth Standards; source={source_url}\n")
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(["day", "minus2sd", "median", "plus2sd"])
        for day, low, median, high in rows:
            writer.writerow([day, f"{low:.3f}", f"{median:.3f}", f"{high:.3f}"])


def main() -> None:
    INPUT.mkdir(parents=True, exist_ok=True)
    for source_name, (target_name, source_url, xlsx_url) in TABLES.items():
        source_path = INPUT / source_name
        if not source_path.exists():
            print(f"download {source_name}")
            urlretrieve(xlsx_url, source_path)
        extract(source_path, OUTPUT / target_name, source_url)
        print(target_name)


if __name__ == "__main__":
    main()
